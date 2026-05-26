"""
Генерация расчётной ведомости в формате Excel (openpyxl).

Лист 1 «Начисления»: удержания из зарплаты (ОПВ, ВОСМС, ИПН, алименты, к выплате).
Лист 2 «Расходы работодателя»: ОПВР, СО, ООСМС, СН, итого затрат.

Подписи: Руководитель + Главный бухгалтер.
Сотрудники расчётную ведомость (аналог Т-51) НЕ подписывают —
они расписываются в платёжной ведомости (Т-53) при получении наличных
или подтверждают банковским извещением.

Настройки страницы: A4, альбомная, вписать в 1 страницу по ширине.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# ─── Стили ────────────────────────────────────────────────────────────────────

_THIN = Side(style="thin", color="000000")
_MEDIUM = Side(style="medium", color="000000")

FULL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
BOTTOM_ONLY = Border(bottom=_MEDIUM)

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")   # голубой для шапки
TOTAL_FILL  = PatternFill("solid", fgColor="EBEBEB")   # серый для итогов

TITLE_FONT = Font(bold=True, size=12, name="Calibri")
HDR_FONT   = Font(bold=True, size=9,  name="Calibri")
DATA_FONT  = Font(size=9,           name="Calibri")
TOTAL_FONT = Font(bold=True, size=9,  name="Calibri")
SIG_FONT   = Font(size=10,          name="Calibri")
HINT_FONT  = Font(size=8, name="Calibri", italic=True, color="808080")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
HDR_AL = Alignment(horizontal="center", vertical="center", wrap_text=True)

NUM_FMT = "# ##0"      # тысячи через пробел (казахстанский формат)


# ─── Вспомогательные функции ──────────────────────────────────────────────────

def _c(ws, row, col, value="", *, font=None, fill=None, border=None, align=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font   is not None: cell.font   = font
    if fill   is not None: cell.fill   = fill
    if border is not None: cell.border = border
    if align  is not None: cell.alignment = align
    if fmt    is not None: cell.number_format = fmt
    return cell


def _header_row(ws, row, headers, height=38):
    ws.row_dimensions[row].height = height
    for col, text in enumerate(headers, 1):
        _c(ws, row, col, text, font=HDR_FONT, fill=HEADER_FILL, border=FULL_BORDER, align=HDR_AL)


def _page_setup(ws):
    ws.page_setup.paperSize   = 9           # A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.39
    ws.page_margins.right  = 0.39
    ws.page_margins.top    = 0.49
    ws.page_margins.bottom = 0.49
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3


def _title_row(ws, text, last_col, row=1, height=22):
    ws.merge_cells(f"A{row}:{get_column_letter(last_col)}{row}")
    _c(ws, row, 1, text, font=TITLE_FONT, align=CENTER)
    ws.row_dimensions[row].height = height


def _total_row(ws, row, label_text, label_cols, value_start_col, values):
    ws.row_dimensions[row].height = 18
    for col in range(1, label_cols + 1):
        style = dict(font=TOTAL_FONT, fill=TOTAL_FILL, border=FULL_BORDER,
                     align=CENTER if col == 1 else LEFT)
        _c(ws, row, col, label_text if col == 1 else "", **style)
    for i, val in enumerate(values):
        _c(ws, row, value_start_col + i, val,
           font=TOTAL_FONT, fill=TOTAL_FILL, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)


def _sig_section(ws, start_row, last_col):
    """
    Строки с подписями: Руководитель (слева) + Главный бухгалтер (справа).
    Место для подписи — линия подчёркивания.
    """
    ws.row_dimensions[start_row].height = 30
    ws.row_dimensions[start_row + 1].height = 14

    mid = max(last_col // 2, 3)

    # Руководитель
    _c(ws, start_row, 1, "Руководитель:", font=SIG_FONT, align=LEFT)
    for col in range(2, mid):
        ws.cell(row=start_row, column=col).border = BOTTOM_ONLY
    _c(ws, start_row + 1, 2, "(подпись / ФИО)", font=HINT_FONT, align=CENTER)

    # Главный бухгалтер
    _c(ws, start_row, mid + 1, "Главный бухгалтер:", font=SIG_FONT, align=LEFT)
    for col in range(mid + 2, last_col + 1):
        ws.cell(row=start_row, column=col).border = BOTTOM_ONLY
    _c(ws, start_row + 1, mid + 2, "(подпись / ФИО)", font=HINT_FONT, align=CENTER)


# ─── Листы ────────────────────────────────────────────────────────────────────

def _build_accruals_sheet(ws, result, period: str):
    ws.title = "Начисления"
    _page_setup(ws)

    col_widths = [4, 32, 20, 14, 12, 12, 12, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last_col = len(col_widths)

    _title_row(ws, f"РАСЧЁТНАЯ ВЕДОМОСТЬ ЗА {period.upper()}", last_col)
    ws.row_dimensions[2].height = 5  # разделитель

    _header_row(ws, 3, [
        "№", "ФИО", "Должность",
        "Оклад\n(брутто)", "ОПВ\n(10%)", "ВОСМС\n(2%)",
        "ИПН", "Алименты", "К выплате",
    ])

    for i, emp in enumerate(result.employees):
        row = 4 + i
        ws.row_dimensions[row].height = 18
        _c(ws, row, 1, i + 1,          font=DATA_FONT, border=FULL_BORDER, align=CENTER)
        _c(ws, row, 2, emp.name or "", font=DATA_FONT, border=FULL_BORDER, align=LEFT)
        _c(ws, row, 3, emp.position or "", font=DATA_FONT, border=FULL_BORDER, align=LEFT)
        _c(ws, row, 4, emp.gross,       font=DATA_FONT, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)
        _c(ws, row, 5, emp.opv,         font=DATA_FONT, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)
        _c(ws, row, 6, emp.osms_employee, font=DATA_FONT, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)
        _c(ws, row, 7, emp.ipn,         font=DATA_FONT, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)
        _c(ws, row, 8, emp.alimony or 0, font=DATA_FONT, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)
        _c(ws, row, 9, emp.to_pay,      font=DATA_FONT, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)

    tot_row = 4 + len(result.employees)
    _total_row(ws, tot_row, "ИТОГО", 3, 4, [
        result.totals.gross, result.totals.opv, result.totals.osms_employee,
        result.totals.ipn, result.totals.alimony, result.totals.to_pay,
    ])

    _sig_section(ws, tot_row + 2, last_col)
    ws.print_area = f"A1:{get_column_letter(last_col)}{tot_row + 3}"


def _build_employer_sheet(ws, result, period: str):
    ws.title = "Расходы работодателя"
    _page_setup(ws)

    col_widths = [4, 32, 14, 14, 12, 12, 12, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last_col = len(col_widths)

    _title_row(ws, f"РАСХОДЫ РАБОТОДАТЕЛЯ ЗА {period.upper()}", last_col)
    ws.row_dimensions[2].height = 5

    _header_row(ws, 3, [
        "№", "ФИО",
        "Оклад\n(брутто)", "ОПВР\n(3.5%)", "СО\n(5%)",
        "ООСМС\n(3%)", "СН\n(6%)", "Итого затрат",
    ])

    for i, emp in enumerate(result.employees):
        row = 4 + i
        ws.row_dimensions[row].height = 18
        _c(ws, row, 1, i + 1,          font=DATA_FONT, border=FULL_BORDER, align=CENTER)
        _c(ws, row, 2, emp.name or "", font=DATA_FONT, border=FULL_BORDER, align=LEFT)
        _c(ws, row, 3, emp.gross,       font=DATA_FONT, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)
        _c(ws, row, 4, emp.opvr,        font=DATA_FONT, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)
        _c(ws, row, 5, emp.so,          font=DATA_FONT, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)
        _c(ws, row, 6, emp.osms_employer, font=DATA_FONT, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)
        _c(ws, row, 7, emp.sn,          font=DATA_FONT, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)
        _c(ws, row, 8, emp.total_cost,  font=DATA_FONT, border=FULL_BORDER, align=RIGHT, fmt=NUM_FMT)

    tot_row = 4 + len(result.employees)
    _total_row(ws, tot_row, "ИТОГО", 2, 3, [
        result.totals.gross, result.totals.opvr, result.totals.so,
        result.totals.osms_employer, result.totals.sn, result.totals.total_cost,
    ])

    _sig_section(ws, tot_row + 2, last_col)
    ws.print_area = f"A1:{get_column_letter(last_col)}{tot_row + 3}"


# ─── Публичный API ────────────────────────────────────────────────────────────

def build_payroll_workbook(result, period: str) -> io.BytesIO:
    wb = Workbook()
    _build_accruals_sheet(wb.active, result, period)
    _build_employer_sheet(wb.create_sheet(), result, period)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
